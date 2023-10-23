# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

#TODO: Excel okuma butonunu yukarıya sabitle onun yerine çizdirme butonunu koy...
#TODO:

import sys

from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QCheckBox, QPushButton, QFileDialog, QGroupBox, QHBoxLayout, QLabel, QTreeWidget, QTreeWidgetItem
from openpyxl import load_workbook
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.widgets import SpanSelector
from matplotlib.backend_bases import PickEvent

def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


class ExcelReaderApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.selected_point = None

        self.initUI()

    def initUI(self):
        self.setWindowTitle("Excel Dosya Okuyucu")
        self.setGeometry(200, 50, 1500, 960)

        # Tam ekran modunu etkinleştir
        #self.showFullScreen()

        # Sağ üst köşede kapatma ve ekranı küçültme butonları ekleyin
        #self.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowTitleHint | Qt.WindowCloseButtonHint | Qt.WindowMinimizeButtonHint)

        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)

        layout = QVBoxLayout()

        self.checkboxes = []
        self.column_names = []

        ## For checkbox_group:
        #self.checkbox_group = QGroupBox("Sütun Başlıkları")
        #self.checkbox_group.setMaximumHeight(50)
        #self.checkbox_group.setMaximumWidth(600)

        self.graph_group = QGroupBox("Grafik")
        graph_layout = QVBoxLayout()
        self.graph_group.setLayout(graph_layout)
        #self.graph_group.setMaximumHeight(700)

        self.open_button = QPushButton("Excel Dosyasını Seç ve Aç")
        self.open_button.clicked.connect(self.openExcelFile)
        self.open_button.setFont(QFont("Arial", 10, QFont.Bold))

        self.plot_button = QPushButton("Grafiği Çizdir")
        self.plot_button.clicked.connect(self.plotGraph)
        self.plot_button.setEnabled(False)
        self.plot_button.setFont(QFont("Arial", 10, QFont.Bold))

        # Ağaç yapısı oluştur
        self.tree_widget = QTreeWidget(self)
        self.tree_widget.setHeaderLabels(["Sütunlar"])
        layout.addWidget(self.tree_widget)
        self.tree_widget.setMaximumHeight(200)
        self.tree_widget.setMinimumHeight(100)

        layout.addWidget(self.open_button)
        ## For checkbox_group:
        #layout.addWidget(self.checkbox_group)
        layout.addWidget(self.graph_group)
        layout.addWidget(self.plot_button)

        central_widget.setLayout(layout)

        # Matplotlib için grafik alanını oluştur
        self.figure = Figure()
        self.canvas = FigureCanvas(self.figure)
        graph_layout.addWidget(self.canvas)


        # Grafiğin zoom işlevselliğini etkinleştiren düğme
        self.zoom_button = QPushButton("Zoom")
        self.zoom_button.clicked.connect(self.enableZoom)
        layout.addWidget(self.zoom_button)
        self.zoom_button.setFont(QFont("Arial", 10, QFont.Bold))

        # Seçili nokta değeri ve etiketi
        #self.selected_point_label = QLabel("Seçili Nokta: ")
        #layout.addWidget(self.selected_point_label)

    def openExcelFile(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly

        excel_file, _ = QFileDialog.getOpenFileName(self, "Excel Dosyasını Seç", "", "Excel Dosyaları (*.xlsx)", options=options)

        if excel_file:
            try:
                self.excel_file = excel_file
                self.workbook = load_workbook(excel_file)
                self.sheet = self.workbook.active

                self.header_row = self.sheet[2]
                for cell in self.header_row:
                    ## For checkbox_group:
                    #checkbox = QCheckBox(cell.value, self)
                    #self.checkboxes.append(checkbox)

                    column_name = cell.value
                    item = QTreeWidgetItem([column_name])
                    item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                    item.setCheckState(0, Qt.Unchecked)
                    self.tree_widget.addTopLevelItem(item)
                    self.column_names.append(column_name)

                ## For checkbox_group:
                #self.updateUI()
                self.plot_button.setEnabled(True)
                print(self.column_names)
            except Exception as e:
                print("Hata:", e)

    def updateUI(self):
        checkbox_layout = QHBoxLayout(self.checkbox_group)
        for checkbox in self.checkboxes:
            checkbox_layout.addWidget(checkbox)

    def getCheckedItems(self):
        items = []
        for i in range(self.tree_widget.topLevelItemCount()):
            item = self.tree_widget.topLevelItem(i)
            if item.checkState(0) == Qt.Checked:
                items.append(item)
        return items

    def plotGraph(self):
        selected_column = None
        self.selected_columns = [item.text(0) for item in self.getCheckedItems()]


        if self.selected_columns:
            self.figure.clear()
            ax = self.figure.add_subplot(111)
            ax.set_title(f"Sütunu Veri Grafiği")
            ax.set_xlabel("Veri Noktası")
            ax.set_ylabel("Değer")
            ax.grid(True)

            for selected_column in self.selected_columns:
                column_index = self.column_names.index(selected_column)
                data = [cell[0].value for cell in
                        self.sheet.iter_rows(min_row=3, max_row=self.sheet.max_row, min_col=column_index + 1,
                                             max_col=column_index + 1)]
                ax.plot(data, label=selected_column)

            # Grafiğe zoom işlevselliği ekleyin
            print("111")
            #self.span = SpanSelector(ax, self.onselect(1,3), 'horizontal', useblit=True, rectprops=dict(alpha=0.5, facecolor='red'))

            self.selected_point = None
            ax.legend()
            self.canvas.draw()

    def enableZoom(self):
        # Zoom düğmesine tıklanınca zoom işlevselliğini etkinleştirin
        self.span.set_active(True)

    def onselect(self, xmin, xmax):
        print(xmin)
        print(xmax)
        print("aaa")
        if self.selected_point:
            print(self.selected_point)
            self.selected_point.remove()
        print("bbb")
        x = (xmin + xmax) / 2
        print(x)
        y = self.get_y_value(x)
        print(y)
        self.selected_point = self.canvas.figure.gca().plot(x, y, 'ro')[0]

        self.selected_point_label.setText(f"Seçili Nokta: x={x:.2f}, y={y:.2f}")
        print("ccc")
        self.canvas.draw()
        print("ddd")
    def get_y_value(self, x):
        # Seçilen x değerine karşılık gelen y değerini hesaplayın
        selected_column = None
        for checkbox in self.checkboxes:
            if checkbox.isChecked():
                selected_column = checkbox.text()
                break

        if selected_column:
            column_index = self.column_names.index(selected_column)
            data = [cell[0].value for cell in
                    self.sheet.iter_rows(min_row=3, max_row=self.sheet.max_row, min_col=column_index + 1,
                                         max_col=column_index + 1)]
            x_values = [i for i in range(len(data))]
            y_values = data
            return y_values[x_values.index(min(x_values, key=lambda val: abs(val - x)))]
        return None


def main():
    app = QApplication(sys.argv)
    window = ExcelReaderApp()
    window.show()
    sys.exit(app.exec_())

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/